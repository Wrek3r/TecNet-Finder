import re
import math
import tkinter as tk
import ipaddress
from tkinter import ttk, font, scrolledtext, messagebox, filedialog
import openpyxl

# --- LÉXICO ---
class VLSMLexer:
    def __init__(self):
        self.tokens = [
            (r'\bIP\b',     'IP'),
            (r'\bMASK\b',   'MASK'),
            (r'\bHOSTS\b',  'HOSTS'),
            (r'\bNAME\b',   'NAME'),
            (r'[0-9]+\.[0-9]+\.[0-9]+\.[0-9]+', 'IP_ADDRESS'),
            (r'/\d+',       'SUBNET_MASK'),
            (r'\d+',        'NUMBER'),
            (r'[A-Za-z_][A-Za-z0-9_]*', 'IDENTIFIER'),
            (r',',          'COMMA'),
            (r'\s+',        None),
        ]
    def tokenize(self, code):
        tokens = []
        errors = []
        line_num = 1
        col_num = 0
        last_type = None
        while code:
            match = None
            for pattern, token_type in self.tokens:
                regex = re.compile(pattern)
                match = regex.match(code)
                if match:
                    text = match.group(0)
                    lines = text.split("\n")
                    if len(lines) > 1:
                        line_num += len(lines) - 1
                        col_num = len(lines[-1])
                    else:
                        col_num += len(text)
                    code = code[len(text):]
                    if token_type:
                        start_col = col_num - len(text)
                        if token_type == 'IDENTIFIER':
                            if last_type == 'NAME':
                                tokens.append((token_type, text, line_num, start_col))
                                last_type = token_type
                            else:
                                errors.append(
                                    f"Token no reconocido en la línea {line_num}, posición {start_col}: '{text}'"
                                )
                        else:
                            tokens.append((token_type, text, line_num, start_col))
                            last_type = token_type
                    break
            if not match:
                error_fragment = code.split()[0] if code.split() else code[0]
                errors.append(
                    f"Token no reconocido en la línea {line_num}, posición {col_num}: '{error_fragment}'"
                )
                code = code[len(error_fragment):]
                col_num += len(error_fragment)
        return tokens, errors


# --- SINTÁCTICO ---
class VLSMParser:
    def __init__(self, tokens):
        self.tokens = tokens
        self.pos = 0
        self.errors = []

    def parse(self):
        results = []
        while self.pos < len(self.tokens):
            try:
                results.append(self.parse_block())
            except SyntaxError as e:
                self.errors.append(str(e))
                self.synchronize()
        return results

    def synchronize(self):
        while self.pos < len(self.tokens) and self.tokens[self.pos][0] != 'IP':
            self.pos += 1

    def parse_block(self):
        self.expect('IP')
        ip_address = self.expect('IP_ADDRESS')
        self.expect('MASK')
        subnet_mask = self.expect('SUBNET_MASK')
        self.expect('HOSTS')
        if self.pos >= len(self.tokens) or self.tokens[self.pos][0] != 'NUMBER':
            token = self.tokens[self.pos] if self.pos < len(self.tokens) else (None, None, '?', '?')
            raise SyntaxError(
                f"Se esperaba al menos un NUMBER para HOSTS pero se encontró "
                f"'{token[0]}' en línea {token[2]}, posición {token[3]}"
            )
        num_hosts = self.parse_hosts()
        name = None
        if self.pos < len(self.tokens) and self.tokens[self.pos][0] == 'NAME':
            self.pos += 1
            if self.pos < len(self.tokens) and self.tokens[self.pos][0] == 'IDENTIFIER':
                name = self.tokens[self.pos][1]
                self.pos += 1
        return {
            'ip_address': ip_address,
            'subnet_mask': subnet_mask,
            'num_hosts': num_hosts,
            'name': name
        }

    def parse_hosts(self):
        hosts = []
        while self.pos < len(self.tokens):
            token = self.tokens[self.pos]
            if token[0] == 'NUMBER':
                hosts.append(int(token[1]))
                self.pos += 1
            elif token[0] == 'COMMA':
                self.pos += 1
            else:
                break
        return hosts

    def expect(self, token_type):
        if self.pos < len(self.tokens):
            token = self.tokens[self.pos]
            if token[0] == token_type:
                self.pos += 1
                return token[1]
            else:
                raise SyntaxError(
                    f"Se esperaba {token_type} pero se encontró '{token[0]}' "
                    f"en línea {token[2]}, posición {token[3]}"
                )
        else:
            raise SyntaxError(f"Se esperaba {token_type} pero no hay más tokens")


# --- SEMÁNTICO ---
class VLSMSemanticAnalyzer:
    def __init__(self):
        self.errors = []

    def analyze(self, blocks):
        """Valida cada bloque semánticamente antes del cálculo VLSM."""
        valid_blocks = []
        for block in blocks:
            if self._validate_block(block):
                valid_blocks.append(block)
        return valid_blocks

    def _validate_block(self, block):
        ip_str    = block['ip_address']
        mask_str  = block['subnet_mask']
        hosts     = block['num_hosts']
        name      = block.get('name', None)
        ok = True

        # 1. Validar dirección IP
        try:
            ip_obj = ipaddress.IPv4Address(ip_str)
        except ValueError:
            self.errors.append(f"IP inválida: '{ip_str}'")
            return False

        # 2. Validar prefijo CIDR
        try:
            cidr = int(mask_str[1:])  # quita el '/'
            if not (0 <= cidr <= 32):
                raise ValueError
        except ValueError:
            self.errors.append(f"Máscara inválida: '{mask_str}'. Debe estar entre /0 y /32.")
            return False

        # 3. Validar que la IP sea dirección de red (host bits = 0)
        try:
            network = ipaddress.IPv4Network(f"{ip_str}{mask_str}", strict=True)
        except ValueError:
            self.errors.append(
                f"'{ip_str}' no es una dirección de red válida para la máscara '{mask_str}'. "
                f"¿Quisiste decir '{ipaddress.IPv4Network(f'{ip_str}{mask_str}', strict=False).network_address}'?"
            )
            ok = False

        # 4. Validar que haya al menos un host
        if not hosts:
            self.errors.append(f"La red '{ip_str}' no tiene hosts definidos.")
            return False

        # 5. Validar valores individuales de hosts
        for h in hosts:
            if h <= 0:
                self.errors.append(f"El número de hosts debe ser mayor a 0 (se encontró {h} en red '{ip_str}').")
                ok = False
            if h > 2**30:
                self.errors.append(f"El número de hosts {h} es demasiado grande en red '{ip_str}'.")
                ok = False

        # 6. Validar que todos los hosts caben dentro de la red base
        if ok:
            try:
                network = ipaddress.IPv4Network(f"{ip_str}{mask_str}", strict=False)
                total_disponibles = network.num_addresses - 2  # sin red ni broadcast
                total_requerido = sum(2 ** math.ceil(math.log2(h + 2)) for h in hosts)
                if total_requerido > network.num_addresses:
                    self.errors.append(
                        f"Los hosts solicitados ({total_requerido} direcciones necesarias) "
                        f"no caben en la red {ip_str}{mask_str} "
                        f"({network.num_addresses} direcciones disponibles)."
                    )
                    ok = False
            except Exception:
                pass

        # 7. Validar nombre de red (si existe)
        if name and not re.match(r'^[A-Za-z_][A-Za-z0-9_]*$', name):
            self.errors.append(f"Nombre de red inválido: '{name}'. Solo letras, números y guion bajo.")
            ok = False

        return ok


# --- CÁLCULO VLSM ---
def calculate_vlsm(ip_address, subnet_mask, num_hosts_list, nombre_red=None):
    results = []
    current_ip = int(ipaddress.IPv4Address(ip_address))
    sorted_hosts = sorted(num_hosts_list, reverse=True)

    for num_hosts in sorted_hosts:
        bits_necesarios = math.ceil(math.log2(num_hosts + 2))
        new_cidr = 32 - bits_necesarios
        block_size = 2 ** bits_necesarios

        network_address   = ipaddress.IPv4Address(current_ip)
        broadcast_address = ipaddress.IPv4Address(current_ip + block_size - 1)
        first_usable_ip   = ipaddress.IPv4Address(current_ip + 1)
        last_usable_ip    = ipaddress.IPv4Address(current_ip + block_size - 2)
        decimal_mask      = str(ipaddress.IPv4Network(f"0.0.0.0/{new_cidr}").netmask)

        results.append({
            'hosts_solicitados':      num_hosts,
            'hosts_disponibles':      block_size - 2,
            'direccion_de_red':       str(network_address),
            'nueva_mascara':          f"/{new_cidr}",
            'mascara_decimal':        decimal_mask,
            'primera_ip_utilizable':  str(first_usable_ip),
            'ultima_ip_utilizable':   str(last_usable_ip),
            'direccion_de_broadcast': str(broadcast_address),
            'ip_base':                ip_address,
            'nombre_red':             nombre_red or ip_address,
        })

        current_ip += block_size

    return results


# --- EXPORTAR A EXCEL ---
def export_to_excel(vlsm_data):
    if not vlsm_data:
        messagebox.showerror("Error", "No hay datos para exportar.")
        return

    file_path = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Archivos de Excel", "*.xlsx")]
    )
    if not file_path:
        return

    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)

    # Agrupar por nombre/ip
    grouped = {}
    for subred in vlsm_data:
        key = subred.get('nombre_red') or subred['ip_base']
        grouped.setdefault(key, []).append(subred)

    header_map = {
        'hosts_solicitados':      'Hosts Solicitados',
        'hosts_disponibles':      'Hosts Disponibles',
        'direccion_de_red':       'Dirección de Red',
        'nueva_mascara':          'Máscara CIDR',
        'mascara_decimal':        'Máscara Decimal',
        'primera_ip_utilizable':  'Primera IP Utilizable',
        'ultima_ip_utilizable':   'Última IP Utilizable',
        'direccion_de_broadcast': 'Broadcast',
    }

    for nombre_red, subredes in grouped.items():
        sheet_name = f"Red {nombre_red}"[:31]  # Excel limita a 31 chars
        sheet = workbook.create_sheet(title=sheet_name)

        # Cabeceras
        headers = ['Subred'] + list(header_map.values())
        for col, h in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col, value=h)
            cell.font = openpyxl.styles.Font(bold=True)

        # Filas de datos
        for i, subred in enumerate(subredes, start=2):
            sheet.cell(row=i, column=1, value=i - 1)
            for col, key in enumerate(header_map.keys(), 2):
                sheet.cell(row=i, column=col, value=subred.get(key, ''))

        # Ajustar ancho de columnas
        for col in sheet.columns:
            max_len = max((len(str(c.value)) for c in col if c.value), default=10)
            sheet.column_dimensions[col[0].column_letter].width = max_len + 4

    workbook.save(file_path)
    messagebox.showinfo("Exportar a Excel", f"Datos exportados exitosamente en:\n{file_path}")


# --- NÚMEROS DE LÍNEA ---
class TextLineNumbers(tk.Canvas):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.textwidget = None

    def attach(self, text_widget):
        self.textwidget = text_widget

    def redraw(self, *args):
        self.delete("all")
        i = self.textwidget.index("@0,0")
        while True:
            dline = self.textwidget.dlineinfo(i)
            if dline is None:
                break
            y = dline[1]
            linenum = str(i).split(".")[0]
            self.create_text(2, y, anchor="nw", text=linenum, fill="#888888")
            i = self.textwidget.index(f"{i}+1line")


# --- GUI ---
class VLSMApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Compilador VLSM")
        self.root.configure(bg="#1e1e2e")

        style = ttk.Style()
        style.theme_use('clam')
        style.configure('TLabelframe',       background='#1e1e2e', foreground='#cdd6f4')
        style.configure('TLabelframe.Label', background='#1e1e2e', foreground='#89b4fa', font=('Courier New', 10, 'bold'))
        style.configure('TFrame',            background='#1e1e2e')
        style.configure('TButton',           background='#313244', foreground='#cdd6f4',
                         font=('Courier New', 10), relief='flat', padding=6)
        style.map('TButton', background=[('active', '#89b4fa')], foreground=[('active', '#1e1e2e')])

        mono = font.Font(family="Courier New", size=10)

        # ── ENTRADA ──────────────────────────────────────────────
        input_frame = ttk.LabelFrame(root, text=" Entrada ")
        input_frame.pack(fill=tk.BOTH, padx=12, pady=(12, 4))

        text_frame = tk.Frame(input_frame, bg="#1e1e2e")
        text_frame.pack(fill=tk.BOTH, expand=True)

        self.linenumbers = TextLineNumbers(text_frame, width=32, bg="#181825", bd=0, highlightthickness=0)
        self.linenumbers.pack(side="left", fill="y")

        self.input_text = scrolledtext.ScrolledText(
            text_frame, wrap=tk.WORD, width=70, height=5,
            font=mono, bg="#181825", fg="#cdd6f4",
            insertbackground="#f38ba8", selectbackground="#313244",
            relief="flat", bd=0
        )
        self.input_text.pack(side="left", fill=tk.BOTH, expand=True)
        self.input_text.bind("<KeyRelease>", self.on_key_release)
        self.linenumbers.attach(self.input_text)

        # Placeholder de ayuda
        hint = "Ejemplo:\nIP 192.168.1.0 MASK /24 HOSTS 50,30,10 NAME Oficina\nIP 10.0.0.0 MASK /8 HOSTS 100,200"
        self.input_text.insert("1.0", hint)
        self.input_text.config(fg="#585b70")
        self.input_text.bind("<FocusIn>",  self._clear_hint)
        self._hint_active = True

        # ── BOTONES ───────────────────────────────────────────────
        btn_frame = ttk.Frame(root)
        btn_frame.pack(pady=6)
        ttk.Button(btn_frame, text="▶  Analizar",        command=self.analyze).pack(side=tk.LEFT, padx=6)
        ttk.Button(btn_frame, text="📊  Exportar Excel",  command=self.export_to_excel).pack(side=tk.LEFT, padx=6)
        ttk.Button(btn_frame, text="🗑  Limpiar",         command=self.clear_all).pack(side=tk.LEFT, padx=6)

        # ── PESTAÑAS DE SALIDA ────────────────────────────────────
        notebook = ttk.Notebook(root)
        notebook.pack(fill=tk.BOTH, expand=True, padx=12, pady=(4, 12))
        style.configure('TNotebook',     background='#1e1e2e', borderwidth=0)
        style.configure('TNotebook.Tab', background='#313244', foreground='#cdd6f4',
                         font=('Courier New', 9), padding=[10, 4])
        style.map('TNotebook.Tab', background=[('selected', '#89b4fa')],
                  foreground=[('selected', '#1e1e2e')])

        # Pestaña 1: Tokens
        tab_tokens = ttk.Frame(notebook)
        notebook.add(tab_tokens, text='🔍 Tokens')
        self.token_text = self._make_output(tab_tokens)

        # Pestaña 2: VLSM
        tab_vlsm = ttk.Frame(notebook)
        notebook.add(tab_vlsm, text='🌐 Tabla VLSM')
        self.vlsm_text = self._make_output(tab_vlsm)

        # Pestaña 3: Errores
        tab_errors = ttk.Frame(notebook)
        notebook.add(tab_errors, text='⚠  Errores')
        self.error_text = self._make_output(tab_errors, fg="#f38ba8")
        ttk.Button(tab_errors, text="Borrar Errores", command=self.clear_errors).pack(pady=4)

        self.vlsm_data = None

    # ── HELPERS ───────────────────────────────────────────────────
    def _make_output(self, parent, fg="#cdd6f4"):
        t = scrolledtext.ScrolledText(
            parent, wrap=tk.WORD, width=80, height=12,
            font=font.Font(family="Courier New", size=10),
            bg="#181825", fg=fg,
            insertbackground="#f38ba8", state=tk.DISABLED,
            relief="flat", bd=4
        )
        t.pack(fill=tk.BOTH, expand=True, padx=4, pady=4)
        return t

    def _write(self, widget, text, clear=False):
        widget.config(state=tk.NORMAL)
        if clear:
            widget.delete("1.0", tk.END)
        widget.insert(tk.END, text)
        widget.config(state=tk.DISABLED)

    def _clear_hint(self, event):
        if self._hint_active:
            self.input_text.delete("1.0", tk.END)
            self.input_text.config(fg="#cdd6f4")
            self._hint_active = False

    # ── EVENTOS ───────────────────────────────────────────────────
    def on_key_release(self, event=None):
        self.linenumbers.redraw()
        self.highlight_reserved_words()

    def highlight_reserved_words(self):
        colors = {'IP': '#89b4fa', 'MASK': '#a6e3a1', 'HOSTS': '#fab387', 'NAME': '#f9e2af'}
        for word, color in colors.items():
            tag = f'rw_{word}'
            self.input_text.tag_remove(tag, '1.0', tk.END)
            idx = '1.0'
            while True:
                idx = self.input_text.search(r'\b' + word + r'\b', idx, tk.END, regexp=True)
                if not idx:
                    break
                end = f"{idx}+{len(word)}c"
                self.input_text.tag_add(tag, idx, end)
                self.input_text.tag_config(tag, foreground=color, font=font.Font(family="Courier New", size=10, weight="bold"))
                idx = end

    # ── ANÁLISIS ──────────────────────────────────────────────────
    def analyze(self):
        if self._hint_active:
            messagebox.showwarning("Advertencia", "Introduce un código para analizar.")
            return

        code = self.input_text.get("1.0", tk.END).strip()
        if not code:
            messagebox.showwarning("Advertencia", "El área de entrada está vacía.")
            return

        self.clear_errors()
        self._write(self.token_text, "", clear=True)
        self._write(self.vlsm_text,  "", clear=True)
        self.vlsm_data = None

        # ── 1. LÉXICO ──
        lexer = VLSMLexer()
        tokens, lex_errors = lexer.tokenize(code)

        token_out = "=== ANÁLISIS LÉXICO ===\n"
        if lex_errors:
            self._write(self.error_text, "=== ERRORES LÉXICOS ===\n")
            for e in lex_errors:
                self._write(self.error_text, f"  {e}\n")
            token_out += f"  {len(lex_errors)} error(es) léxico(s) encontrado(s).\n"
        else:
            for t in tokens:
                ttype, val, line, col = t
                token_out += f"  [{ttype}]  '{val}'  →  línea {line}, col {col}\n"

        self._write(self.token_text, token_out, clear=True)

        if lex_errors:
            return  # No continuar si hay errores léxicos

        # ── 2. SINTÁCTICO ──
        parser = VLSMParser(tokens)
        blocks = parser.parse()

        if parser.errors:
            self._write(self.error_text, "\n=== ERRORES SINTÁCTICOS ===\n")
            for e in parser.errors:
                self._write(self.error_text, f"  {e}\n")

        if not blocks:
            self._write(self.vlsm_text, "No se encontraron bloques válidos.\n", clear=True)
            return

        # ── 3. SEMÁNTICO ──
        semantic = VLSMSemanticAnalyzer()
        valid_blocks = semantic.analyze(blocks)

        if semantic.errors:
            self._write(self.error_text, "\n=== ERRORES SEMÁNTICOS ===\n")
            for e in semantic.errors:
                self._write(self.error_text, f"  {e}\n")

        if not valid_blocks:
            self._write(self.vlsm_text, "No hay bloques semánticamente válidos para calcular.\n", clear=True)
            return

        # ── 4. CÁLCULO VLSM ──
        all_results = []
        for block in valid_blocks:
            try:
                result = calculate_vlsm(
                    block['ip_address'],
                    block['subnet_mask'],
                    block['num_hosts'],
                    nombre_red=block.get('name')
                )
                all_results.extend(result)
            except Exception as ex:
                self._write(self.error_text, f"\n  Error al calcular VLSM para {block['ip_address']}: {ex}\n")

        self.vlsm_data = all_results

        # ── 5. MOSTRAR TABLA VLSM ──
        vlsm_out = "=== TABLA VLSM ===\n\n"
        grouped = {}
        for r in all_results:
            key = r.get('nombre_red') or r['ip_base']
            grouped.setdefault(key, []).append(r)

        for nombre, subredes in grouped.items():
            vlsm_out += f"  Red: {nombre}\n"
            vlsm_out += f"  {'─'*70}\n"
            vlsm_out += (
                f"  {'#':<4} {'Hosts Solic.':<14} {'Hosts Disp.':<13} "
                f"{'Dirección Red':<18} {'CIDR':<8} {'Máscara':<18} "
                f"{'Primera IP':<16} {'Última IP':<16} {'Broadcast'}\n"
            )
            vlsm_out += f"  {'─'*70}\n"
            for i, s in enumerate(subredes, 1):
                vlsm_out += (
                    f"  {i:<4} {s['hosts_solicitados']:<14} {s['hosts_disponibles']:<13} "
                    f"{s['direccion_de_red']:<18} {s['nueva_mascara']:<8} {s['mascara_decimal']:<18} "
                    f"{s['primera_ip_utilizable']:<16} {s['ultima_ip_utilizable']:<16} {s['direccion_de_broadcast']}\n"
                )
            vlsm_out += "\n"

        self._write(self.vlsm_text, vlsm_out, clear=True)

        if not parser.errors and not semantic.errors:
            messagebox.showinfo("Análisis completo", f"✅ {len(all_results)} subred(es) calculada(s) correctamente.")

    # ── LIMPIAR ───────────────────────────────────────────────────
    def clear_errors(self):
        self._write(self.error_text, "", clear=True)

    def clear_all(self):
        self.clear_errors()
        self._write(self.token_text, "", clear=True)
        self._write(self.vlsm_text,  "", clear=True)
        self.input_text.config(state=tk.NORMAL)
        self.input_text.delete("1.0", tk.END)
        self.vlsm_data = None

    # ── EXPORTAR ─────────────────────────────────────────────────
    def export_to_excel(self):
        if self.vlsm_data:
            export_to_excel(self.vlsm_data)
        else:
            messagebox.showerror("Error", "Primero analiza un código válido para exportar.")


if __name__ == "__main__":
    root = tk.Tk()
    root.minsize(800, 600)
    app = VLSMApp(root)
    root.mainloop()
