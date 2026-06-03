# Compilador.py
# Implementación del compilador para TecNet Finder.
# Incluye el análisis léxico, sintáctico y el cálculo VLSM.
# También tiene funciones para exportar resultados a Excel y mostrar números de línea.

# ----------------------------------- IMPORTS -----------------------------------
# Se importan las librerías necesarias para el funcionamiento del compilador.
import re
import math
import tkinter as tk
import ipaddress
from tkinter import ttk, font, scrolledtext, messagebox, filedialog
from collections import Counter
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from PIL import Image, ImageDraw, ImageFont # type: ignore

# Descripciones breves para mostrar errores más claros.
TOKEN_DESCRIPTIONS = {
    'IP': 'palabra reservada IP',
    'MASK': 'palabra reservada MASK',
    'HOSTS': 'palabra reservada HOSTS',
    'NAME': 'palabra reservada NAME',
    'IP_ADDRESS': 'dirección IP completa',
    'SUBNET_MASK': 'máscara CIDR',
    'NUMBER': 'número entero',
    'IDENTIFIER': 'identificador o nombre de red',
    'COMMA': 'coma separadora',
    'DOT': 'punto',
}

# Devuelve una descripción del token esperado.
# Para IP_ADDRESS y SUBNET_MASK intenta usar como ejemplo lo que el usuario escribió.
def describe_expected_token(token_type, current_token=None, tokens=None, pos=0):
    if token_type == 'SUBNET_MASK' and current_token and current_token[0] == 'NUMBER':
        return f"máscara CIDR, por ejemplo /{current_token[1]}"

    if token_type == 'IP_ADDRESS' and tokens is not None:
        parts = []
        i = pos

        while i < len(tokens) and tokens[i][0] in ('NUMBER', 'DOT'):
            parts.append(tokens[i][1])
            i += 1

        partial_ip = ''.join(parts)

        if partial_ip:
            if partial_ip.endswith('.'):
                partial_ip += '0'

            while partial_ip.count('.') < 3:
                partial_ip += '.0'

            octets = partial_ip.split('.')

            while len(octets) < 4:
                octets.append('0')

            fixed_octets = []
            for octet in octets[:4]:
                if octet.isdigit() and 0 <= int(octet) <= 255:
                    fixed_octets.append(octet)
                else:
                    fixed_octets.append('0')

            return f"dirección IP completa, por ejemplo {'.'.join(fixed_octets)}"

        return "dirección IP completa, por ejemplo 192.168.1.0"

    return TOKEN_DESCRIPTIONS.get(token_type, token_type)

# ----------------------------------- LÉXICO -----------------------------------
# Analizador léxico de TecNet Finder.
# Recorre el texto ingresado y lo convierte en tokens.
class VLSMLexer:
    # Inicializa las expresiones regulares usadas por el lexer.
    # El orden de los patrones define la prioridad de reconocimiento.
    def __init__(self):
        # Patrón para validar un octeto IPv4.
        # Acepta valores desde 0 hasta 255.
        octet = r'(?:25[0-5]|2[0-4][0-9]|1[0-9]{2}|[1-9]?[0-9])'

        # Patrón para validar una dirección IPv4 completa.
        ip_regex = rf'\b{octet}\.{octet}\.{octet}\.{octet}\b'

        # Lista de patrones léxicos.
        # Cada elemento contiene una expresión regular y el tipo de token que genera.
        # El orden es importante porque las palabras reservadas se revisan antes que IDENTIFIER.
        self.tokens = [
            # Palabras reservadas.
            # Se usa lookahead (?=...) para reconocerlas también si vienen pegadas
            # al valor esperado, por ejemplo IP192 o MASK/24.
            (r'\bIP(?=\s|\d|$)', 'IP'),
            (r'\bMASK(?=\s|/|$)', 'MASK'),
            (r'\bHOSTS(?=\s|\d|,|$)', 'HOSTS'),
            (r'\bNAME\b', 'NAME'),

            # Primero se intenta reconocer una IP completa.
            # Si la IP está incompleta, se reconocerán sus partes como NUMBER y DOT.
            (ip_regex, 'IP_ADDRESS'),

            (r'/\d+', 'SUBNET_MASK'),

            # Token agregado para reconocer el punto como parte válida del lenguaje.
            # Esto permite que una IP incompleta se mande al análisis sintáctico.
            (r'\.', 'DOT'),

            (r'\d+', 'NUMBER'),
            (r'[A-Za-z_][A-Za-z0-9_]*', 'IDENTIFIER'),
            (r',', 'COMMA'),
            (r'\s+', None),
        ]

    # Convierte el texto de entrada en tokens.
    # Si una parte no coincide con ningún patrón, se registra como error léxico.
    def tokenize(self, code):
        tokens = []
        errores = []
        line_num = 1
        col_num = 0
        last_type = None

        while code:
            match = None
            for pattern, token_type in self.tokens:
                regex = re.compile(pattern)
                match = regex.match(code)
                if match:
                    text = match.group(0)
                    lines = text.split("\n")
                    if len(lines) > 1:
                        line_num += len(lines) - 1
                        col_num = len(lines[-1])
                    else:
                        col_num += len(text)
                    code = code[len(text):]
                    if token_type:
                        start_col = col_num - len(text)
                        if token_type == 'IDENTIFIER':
                            if last_type == 'NAME':
                                tokens.append((token_type, text, line_num, start_col))
                                last_type = token_type
                            else:
                                errores.append(
                                    f"Identificador fuera de lugar en la línea {line_num}, posición {start_col}: '{text}'. "
                                    f"Los identificadores solo son válidos después de NAME."
                                )
                        else:
                            tokens.append((token_type, text, line_num, start_col))
                            last_type = token_type
                    break
            if not match:
                error_fragment = code.split()[0] if code.split() else code[0]
                errores.append(
                    f"Símbolo o fragmento no reconocido en la línea {line_num}, posición {col_num}: '{error_fragment}'. "
                    f"Revise que la entrada use palabras reservadas válidas como IP, MASK, HOSTS o NAME."
                )
                code = code[len(error_fragment):]
                col_num += len(error_fragment)
        return tokens, errores


# ----------------------------------- SINTÁCTICO -----------------------------------
# Esta clase recibe los tokens generados por el lexer
# y verifica que aparezcan en el orden correcto.
class VLSMParser:
    # Inicializa el parser con la lista de tokens recibida.
    # También prepara la posición actual y la lista de errores.
    def __init__(self, tokens):
        self.tokens = tokens
        self.pos = 0
        self.errors = []

    # Procesa todos los bloques de entrada.
    # Usa modo pánico para continuar después de errores sintácticos.
    def parse(self):
        results = []
        seen_blocks = set()

        while self.pos < len(self.tokens):
            start_pos = self.pos

            try:
                errors_before = len(self.errors)
                block = self.parse_block()
                block_has_errors = len(self.errors) > errors_before

                signature = (
                    block["ip_address"],
                    block["subnet_mask"],
                    tuple(block["num_hosts"]),
                    block["name"]
                )

                if signature in seen_blocks:
                    line = self.tokens[start_pos][2]
                    self.errors.append(
                        f"Error sintáctico: instrucción duplicada en línea {line}. "
                        f"La red con IP {block['ip_address']}, máscara {block['subnet_mask']} "
                        f"y nombre {block['name']} ya fue declarada anteriormente."
                    )

                elif not block_has_errors:
                    seen_blocks.add(signature)
                    results.append(block)

            except SyntaxError as e:
                self.errors.append(str(e))
                self.synchronize(start_pos)

            if self.pos == start_pos:
                self.pos += 1

        return results
    
    # Modo pánico: avanza hasta encontrar el inicio de otra instrucción.
    # En este lenguaje, una nueva instrucción inicia con IP.
    def synchronize(self, start_pos=None):
        error_line = None

        if start_pos is not None and start_pos < len(self.tokens):
            error_line = self.tokens[start_pos][2]

        while self.pos < len(self.tokens):
            token_type, token_value, line, col = self.tokens[self.pos]

            if token_type == "IP":
                if error_line is None or line > error_line:
                    return

            self.pos += 1

    # Valida una instrucción completa con la estructura:
    # IP <IP_ADDRESS> MASK <SUBNET_MASK> HOSTS <lista_hosts> [NAME <IDENTIFIER>]
    def parse_block(self):
        source_line = self.tokens[self.pos][2]

        self.expect('IP')

        # Después de IP debe venir una dirección IP completa.
        # Si aparecen NUMBER, DOT u otros tokens, el error es sintáctico,
        # porque los tokens existen, pero no forman la estructura esperada.
        if self.pos >= len(self.tokens) or self.tokens[self.pos][0] != 'IP_ADDRESS':
            ejemplo_ip = describe_expected_token('IP_ADDRESS', None, self.tokens, self.pos)

            if self.pos >= len(self.tokens):
                raise SyntaxError(
                    f"Error sintáctico: dirección IP incompleta o mal formada. "
                    f"Después de IP se esperaba IP_ADDRESS ({ejemplo_ip}), "
                    f"pero se llegó al fin de la entrada."
                )

            token = self.tokens[self.pos]

            raise SyntaxError(
                f"Error sintáctico: dirección IP incompleta o mal formada. "
                f"Después de IP se esperaba IP_ADDRESS ({ejemplo_ip}), "
                f"pero se encontró {token[0]} ('{token[1]}') "
                f"en línea {token[2]}, posición {token[3]}."
            )

        ip_address = self.expect('IP_ADDRESS')
        self.expect('MASK')
        subnet_mask = self.expect('SUBNET_MASK')
        
        self.expect('HOSTS')
        if self.pos >= len(self.tokens) or self.tokens[self.pos][0] != 'NUMBER':
            if self.pos >= len(self.tokens):
                raise SyntaxError(
                    f"Error sintáctico: después de HOSTS se esperaba al menos un NUMBER "
                    f"con la cantidad de hosts, por ejemplo 50, "
                    f"pero se llegó al fin de la entrada."
                )

            token = self.tokens[self.pos]

            raise SyntaxError(
                f"Error sintáctico: después de HOSTS se esperaba al menos un NUMBER "
                f"con la cantidad de hosts, por ejemplo 50, "
                f"pero se encontró {token[0]} ('{token[1]}') "
                f"en línea {token[2]}, posición {token[3]}."
            )
        num_hosts = self.parse_hosts()
        name = None

        # El nombre de la red es opcional, pero si aparece NAME,
        # obligatoriamente debe venir un IDENTIFIER después.
        if self.pos < len(self.tokens) and self.tokens[self.pos][0] == 'NAME':
            name_token = self.tokens[self.pos]
            self.pos += 1

            if self.pos >= len(self.tokens):
                raise SyntaxError(
                    f"Error sintáctico: después de NAME se esperaba IDENTIFIER "
                    f"con el nombre de la red, pero se llegó al fin de la entrada "
                    f"en línea {name_token[2]}, posición {name_token[3]}."
                )

            token = self.tokens[self.pos]

            if token[0] == 'IP' and token[2] > name_token[2]:
                raise SyntaxError(
                    f"Error sintáctico: después de NAME se esperaba IDENTIFIER "
                    f"con el nombre de la red, pero la instrucción terminó sin indicar el nombre "
                    f"antes de iniciar otra instrucción en línea {token[2]}, posición {token[3]}."
                )

            if token[0] != 'IDENTIFIER':
                raise SyntaxError(
                    f"Error sintáctico: después de NAME se esperaba IDENTIFIER "
                    f"con el nombre de la red, pero se encontró {token[0]} ('{token[1]}') "
                    f"en línea {token[2]}, posición {token[3]}."
                )

            name = self.tokens[self.pos][1]
            self.pos += 1
        return {
            'ip_address': ip_address,
            'subnet_mask': subnet_mask,
            'num_hosts': num_hosts,
            'name': name,
            'source_line': source_line
        }

    # Procesa la lista de hosts separados por comas.
    # Si encuentra errores, intenta recuperarse para seguir revisando el bloque.
    def parse_hosts(self):
        hosts = []

        # El primer valor de HOSTS debe ser un número.
        token = self.tokens[self.pos]
        hosts.append(int(token[1]))
        self.pos += 1

        # Después de una coma siempre debe venir otro número.
        while self.pos < len(self.tokens) and self.tokens[self.pos][0] == 'COMMA':
            comma_token = self.tokens[self.pos]
            self.pos += 1

            if self.pos >= len(self.tokens):
                self.errors.append(
                    f"Error sintáctico: después de la coma se esperaba NUMBER "
                    f"con otra cantidad de hosts, pero no hay más tokens "
                    f"en línea {comma_token[2]}, posición {comma_token[3]}."
                )
                return hosts

            token = self.tokens[self.pos]

            if token[0] != 'NUMBER':
                self.errors.append(
                    f"Error sintáctico: después de la coma se esperaba NUMBER "
                    f"con otra cantidad de hosts, pero se encontró {token[0]} ('{token[1]}') "
                    f"en línea {token[2]}, posición {token[3]}."
                )

                # Recuperación local:
                # Avanza hasta encontrar NUMBER, NAME o el inicio de otra instrucción IP.
                while (
                    self.pos < len(self.tokens)
                    and self.tokens[self.pos][0] not in ('NUMBER', 'NAME', 'IP')
                ):
                    self.pos += 1

                # Si se recupera encontrando otro número, lo toma y continúa.
                if self.pos < len(self.tokens) and self.tokens[self.pos][0] == 'NUMBER':
                    hosts.append(int(self.tokens[self.pos][1]))
                    self.pos += 1
                    continue

                # Si encuentra NAME o IP, deja que parse_block o parse controle lo siguiente.
                return hosts

            hosts.append(int(token[1]))
            self.pos += 1

        return hosts
    
    # Verifica que el token actual sea del tipo esperado.
    # Si coincide, avanza; si no coincide, genera un error sintáctico.
    def expect(self, token_type):
        if self.pos < len(self.tokens):
            token = self.tokens[self.pos]

            if token[0] == token_type:
                self.pos += 1
                return token[1]
            else:
                esperado = describe_expected_token(token_type, token, self.tokens, self.pos)

                raise SyntaxError(
                    f"Error sintáctico: se esperaba {token_type} ({esperado}), "
                    f"pero se encontró {token[0]} ('{token[1]}') "
                    f"en línea {token[2]}, posición {token[3]}."
                )
        else:
            esperado = describe_expected_token(token_type)
            raise SyntaxError(
                f"Error sintáctico: se esperaba {token_type} ({esperado}), "
                f"pero no hay más tokens."
            )

    # Genera la estructura de árbol a partir de los tokens.
    # Devuelve un árbol por cada bloque válido encontrado.
    def parse_with_tree(self):
        tree = []

        while self.pos < len(self.tokens):
            try:
                node = self.parse_block_tree()
                tree.append(node)
            except SyntaxError as e:
                self.errors.append(str(e))
                self.synchronize()

        return tree

    # Crea un nodo del árbol para un token y su valor.
    # Se usa para separar visualmente tipo de token y contenido.
    def token_tree(self, token_type, value):
        return (token_type, [("VALOR", value)])

    # Construye el árbol sintáctico de una instrucción completa.
    # Organiza IP, máscara, hosts y nombre opcional.
    def parse_block_tree(self):
        children = []

        ip_token = self.expect("IP")
        children.append(self.token_tree("IP", ip_token))

        ip_address = self.expect("IP_ADDRESS")
        children.append(self.token_tree("IP_ADDRESS", ip_address))

        mask_token = self.expect("MASK")
        children.append(self.token_tree("MASK", mask_token))

        subnet_mask = self.expect("SUBNET_MASK")
        children.append(self.token_tree("SUBNET_MASK", subnet_mask))

        hosts_token = self.expect("HOSTS")
        children.append(self.token_tree("HOSTS", hosts_token))

        hosts_node = self.parse_hosts_tree()
        children.append(("HOSTS_LIST", hosts_node))

        name = None

        if self.pos < len(self.tokens) and self.tokens[self.pos][0] == "NAME":
            name_token = self.expect("NAME")
            children.append(self.token_tree("NAME", name_token))

            identifier = self.expect("IDENTIFIER")
            children.append(self.token_tree("IDENTIFIER", identifier))

            name = identifier

        if name:
            root_label = f"BloqueRed: {name}"
        else:
            root_label = "BloqueRed"

        return (root_label, children)

    # Construye el subárbol de la lista de hosts.
    # Incluye números y comas reconocidas por el lexer.
    def parse_hosts_tree(self):
        hosts = []

        number = self.expect("NUMBER")
        hosts.append(self.token_tree("NUMBER", number))

        while self.pos < len(self.tokens) and self.tokens[self.pos][0] == "COMMA":
            comma = self.expect("COMMA")
            hosts.append(self.token_tree("COMMA", comma))

            number = self.expect("NUMBER")
            hosts.append(self.token_tree("NUMBER", number))

        return hosts

# ----------------------------------- CÁLCULO VLSM -----------------------------------
# Calcula las subredes VLSM a partir de una IP base,
# una máscara y una lista de hosts requeridos.
def calculate_vlsm(ip_address, subnet_mask, num_hosts_list, nombre_red=None):
    results = []
    current_ip = int(ipaddress.IPv4Address(ip_address))

    # Ordena los hosts de mayor a menor para asignar primero las subredes más grandes.
    sorted_hosts = sorted(num_hosts_list, reverse=True)

    for num_hosts in sorted_hosts:
        # Se suman 2 direcciones: una para red y una para broadcast.
        bits_necesarios = math.ceil(math.log2(num_hosts + 2))
        new_cidr = 32 - bits_necesarios
        block_size = 2 ** bits_necesarios

        network_address   = ipaddress.IPv4Address(current_ip)
        broadcast_address = ipaddress.IPv4Address(current_ip + block_size - 1)
        first_usable_ip   = ipaddress.IPv4Address(current_ip + 1)
        last_usable_ip    = ipaddress.IPv4Address(current_ip + block_size - 2)
        decimal_mask      = str(ipaddress.IPv4Network(f"0.0.0.0/{new_cidr}").netmask)

        # Guarda los datos calculados de la subred actual.
        results.append({
            'hosts_solicitados':      num_hosts,
            'hosts_disponibles':      block_size - 2,
            'direccion_de_red':       str(network_address),
            'nueva_mascara':          f"/{new_cidr}",
            'mascara_decimal':        decimal_mask,
            'primera_ip_utilizable':  str(first_usable_ip),
            'ultima_ip_utilizable':   str(last_usable_ip),
            'direccion_de_broadcast': str(broadcast_address),
            'ip_base':                ip_address,
            'nombre_red':             nombre_red or ip_address,
        })

        # Avanza a la siguiente dirección disponible.
        current_ip += block_size

    return results

# Limpia nombres para usarlos en archivos u hojas de Excel.
# Reemplaza caracteres no válidos por guiones bajos.
def clean_name_for_file(name, default="Red"):
    name = str(name).strip()

    if not name:
        name = default

    name = re.sub(r'[^A-Za-z0-9_-]', '_', name)

    if not name:
        name = default

    return name

# Exporta los datos de VLSM a un archivo Excel.
# Agrupa por nombre de red o IP base, y crea una hoja para cada grupo.
def export_to_excel(vlsm_data):
    if not vlsm_data:
        messagebox.showerror("Error", "No hay datos para exportar.")
        return

    grouped = {}

    for subred in vlsm_data:
        key = subred.get("nombre_red") or subred.get("ip_base", "Red")
        grouped.setdefault(key, []).append(subred)

    if len(grouped) == 1:
        first_name = list(grouped.keys())[0]
        nombre_archivo = f"VLSM_{clean_name_for_file(first_name)}.xlsx"
    else:
        nombre_archivo = "VLSM_TecNetFinder.xlsx"

    file_path = filedialog.asksaveasfilename(
        initialfile=nombre_archivo,
        defaultextension=".xlsx",
        filetypes=[("Archivos de Excel", "*.xlsx")]
    )

    if not file_path:
        return

    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)

    used_sheet_names = set()

    def clean_sheet_name(name):
        name = str(name).strip()
        name = re.sub(r'[\[\]\:\*\?\/\\]', '_', name)

        if not name:
            name = "Red"

        name = name[:31]
        original = name
        counter = 1

        while name in used_sheet_names:
            suffix = f"_{counter}"
            name = original[:31 - len(suffix)] + suffix
            counter += 1

        used_sheet_names.add(name)
        return name

    headers = [
        "Subred",
        "Hosts solicitados",
        "Hosts disponibles",
        "Dirección de red",
        "CIDR",
        "Máscara decimal",
        "Primera IP utilizable",
        "Última IP utilizable",
        "Broadcast",
        "Nombre de red"
    ]

    header_fill = PatternFill(
        start_color="313244",
        end_color="313244",
        fill_type="solid"
    )

    header_font = Font(
        color="CDD6F4",
        bold=True
    )

    row_fill_1 = PatternFill(
        start_color="F5F5F5",
        end_color="F5F5F5",
        fill_type="solid"
    )

    row_fill_2 = PatternFill(
        start_color="FFFFFF",
        end_color="FFFFFF",
        fill_type="solid"
    )

    thin_border = Border(
        left=Side(style="thin", color="89B4FA"),
        right=Side(style="thin", color="89B4FA"),
        top=Side(style="thin", color="89B4FA"),
        bottom=Side(style="thin", color="89B4FA")
    )

    center_alignment = Alignment(
        horizontal="center",
        vertical="center"
    )

    for nombre_red, subredes in grouped.items():
        sheet = workbook.create_sheet(title=clean_sheet_name(nombre_red))
        sheet.append(headers)

        for index, subred in enumerate(subredes, start=1):
            sheet.append([
                index,
                subred.get("hosts_solicitados", ""),
                subred.get("hosts_disponibles", ""),
                subred.get("direccion_de_red", ""),
                subred.get("nueva_mascara", ""),
                subred.get("mascara_decimal", ""),
                subred.get("primera_ip_utilizable", ""),
                subred.get("ultima_ip_utilizable", ""),
                subred.get("direccion_de_broadcast", ""),
                subred.get("nombre_red", "")
            ])

        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
            cell.alignment = center_alignment

        for row_index, row in enumerate(sheet.iter_rows(min_row=2), start=2):
            fill = row_fill_1 if row_index % 2 == 0 else row_fill_2

            for cell in row:
                cell.fill = fill
                cell.border = thin_border
                cell.alignment = center_alignment

        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                value = str(cell.value) if cell.value is not None else ""
                max_length = max(max_length, len(value))

            sheet.column_dimensions[column_letter].width = max_length + 3

        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions

    workbook.save(file_path)
    messagebox.showinfo("Exportar a Excel", "Datos exportados correctamente.")

# ----------------------------------- NÚMEROS DE LÍNEA --------------------------------------
# Canvas auxiliar para mostrar números de línea.
# Se sincroniza con el área de entrada del usuario.
class TextLineNumbers(tk.Canvas):
    # Inicializa el canvas de números de línea.
    # La referencia al área de texto se asigna después.
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.textwidget = None

    # Asocia el canvas con el área de texto principal.
    # Permite leer las líneas visibles del editor.
    def attach(self, text_widget):
        self.textwidget = text_widget

    # Redibuja los números de línea visibles.
    # Se actualiza cuando cambia el contenido de entrada.
    def redraw(self, *args):
        self.delete("all")
        i = self.textwidget.index("@0,0")
        while True:
            dline = self.textwidget.dlineinfo(i)
            if dline is None:
                break
            y = dline[1]
            linenum = str(i).split(".")[0]
            self.create_text(2, y, anchor="nw", text=linenum, fill="#888888")
            i = self.textwidget.index(f"{i}+1line")

# Muestra ayuda breve al pasar el cursor sobre un botón.
# Se usa en botones pequeños como pegar, copiar y guardar.
class Tooltip:
    # Asocia el tooltip con un widget específico.
    # También registra los eventos de entrada y salida del cursor.
    def __init__(self, widget, text):
        self.widget = widget
        self.text = text
        self.tooltip_window = None

        self.widget.bind("<Enter>", self.show)
        self.widget.bind("<Leave>", self.hide)

    # Muestra la ventana flotante del tooltip.
    # Ajusta la posición para evitar que salga de la pantalla.
    def show(self, event=None):
        if self.tooltip_window or not self.text:
            return

        self.tooltip_window = tk.Toplevel(self.widget)
        self.tooltip_window.wm_overrideredirect(True)

        label = tk.Label(
            self.tooltip_window,
            text=self.text,
            bg="#313244",
            fg="#cdd6f4",
            relief="solid",
            borderwidth=1,
            font=("Courier New", 9),
            padx=6,
            pady=3
        )
        label.pack()

        self.tooltip_window.update_idletasks()

        screen_width = self.widget.winfo_screenwidth()
        screen_height = self.widget.winfo_screenheight()

        tooltip_width = self.tooltip_window.winfo_reqwidth()
        tooltip_height = self.tooltip_window.winfo_reqheight()

        x = self.widget.winfo_rootx() + 10
        y = self.widget.winfo_rooty() + self.widget.winfo_height() + 5

        if x + tooltip_width > screen_width:
            x = screen_width - tooltip_width - 10

        if y + tooltip_height > screen_height:
            y = self.widget.winfo_rooty() - tooltip_height - 5

        self.tooltip_window.wm_geometry(f"+{x}+{y}")
    
    # Oculta y destruye la ventana del tooltip.
    # Se ejecuta cuando el cursor sale del botón.
    def hide(self, event=None):
        if self.tooltip_window:
            self.tooltip_window.destroy()
            self.tooltip_window = None


# ----------------------------------- INTERFAZ GRÁFICA -----------------------------------
# Clase principal de la interfaz gráfica de TecNet Finder.
# Permite ingresar instrucciones, analizarlas y mostrar tokens, errores y tabla VLSM.
class VLSMApp:
    # Inicializa la ventana principal y los estilos visuales.
    # También crea las pestañas principales de la aplicación.
    def __init__(self, root):
        self.root = root
        self.root.title("TecNet Finder")
        self.root.configure(bg="#1e1e2e")

        style = ttk.Style()
        style.theme_use('clam')

        style.configure(
            'TLabelframe',
            background='#1e1e2e',
            foreground='#cdd6f4'
        )

        style.configure(
            'TLabelframe.Label',
            background='#1e1e2e',
            foreground='#89b4fa',
            font=('Courier New', 10, 'bold')
        )

        style.configure(
            'TFrame',
            background='#1e1e2e'
        )

        style.configure(
            'TButton',
            background='#313244',
            foreground='#cdd6f4',
            font=('Courier New', 10),
            relief='flat',
            padding=6
        )

        style.map(
            'TButton',
            background=[('active', '#89b4fa')],
            foreground=[('active', '#1e1e2e')]
        )

        style.configure(
            'TNotebook',
            background='#1e1e2e',
            borderwidth=0
        )

        style.configure(
            'TNotebook.Tab',
            background='#313244',
            foreground='#cdd6f4',
            font=('Courier New', 9),
            padding=[10, 4]
        )

        style.map(
            'TNotebook.Tab',
            background=[('selected', '#89b4fa')],
            foreground=[('selected', '#1e1e2e')]
        )

        style.configure(
            "Treeview",
            background="#181825",
            fieldbackground="#181825",
            foreground="#cdd6f4",
            rowheight=24,
            borderwidth=0,
            font=("Courier New", 10)
        )

        style.configure(
            "Treeview.Heading",
            background="#313244",
            foreground="#cdd6f4",
            font=("Courier New", 10, "bold")
        )

        style.map(
            "Treeview",
            background=[("selected", "#89b4fa")],
            foreground=[("selected", "#1e1e2e")]
        )

        self.mono = font.Font(family="Courier New", size=10)

        self.vlsm_data = None
        self.tokens = []
        self.valid_blocks = []
        self.tree_tabs = []

        self.hint_text = (
            "Ejemplo:\n"
            "IP 192.168.1.0 MASK /24 HOSTS 50,30,10 NAME Oficina\n"
            "IP 10.0.0.0 MASK /8 HOSTS 100,200"
        )

        self._hint_active = False

        # Notebook principal
        self.notebook = ttk.Notebook(root)
        self.notebook.pack(fill=tk.BOTH, expand=True, padx=12, pady=12)

        # Pestañas principales
        self.tab_io = ttk.Frame(self.notebook)
        self.tab_tables = ttk.Frame(self.notebook)
        self.tab_tree = ttk.Frame(self.notebook)

        self.notebook.add(self.tab_io, text="Entrada / Salida")
        self.notebook.add(self.tab_tables, text="Tablas")
        self.notebook.add(self.tab_tree, text="Árbol")

        # Construcción de cada pestaña principal
        self._build_io_tab(self.tab_io)
        self._build_tables(self.tab_tables)
        self._build_tree(self.tab_tree)

    # Construye la pestaña de entrada y salida.
    # Incluye entrada, botones, análisis, tabla VLSM y errores.
    def _build_io_tab(self, parent):

        # ───────────────────── ENTRADA ─────────────────────
        
        input_frame = ttk.LabelFrame(parent, text=" Entrada ")
        input_frame.pack(fill=tk.BOTH, padx=12, pady=(12, 4))

        text_frame = tk.Frame(input_frame, bg="#1e1e2e")
        text_frame.pack(fill=tk.BOTH, expand=True, padx=4, pady=4)

        # Números de línea a la izquierda del área de texto.
        self.linenumbers = TextLineNumbers(
            text_frame,
            width=32,
            bg="#181825",
            bd=0,
            highlightthickness=0
        )
        self.linenumbers.pack(side="left", fill="y")

        # Área de texto para ingresar las instrucciones.
        self.input_text = scrolledtext.ScrolledText(
            text_frame,
            wrap=tk.WORD,
            width=70,
            height=5,
            font=self.mono,
            bg="#181825",
            fg="#cdd6f4",
            insertbackground="#f38ba8",
            selectbackground="#313244",
            relief="flat",
            bd=0
        )
        self.input_text.pack(side="left", fill=tk.BOTH, expand=True)
        self.input_text.bind("<KeyRelease>", self.on_key_release)
        self.linenumbers.attach(self.input_text)

        paste_button = self._make_icon_button(
            text_frame,
            "📄",
            self.paste_clipboard_to_input,
            "Pegar comando desde portapapeles"
        )
        paste_button.place(relx=1.0, x=-23, y=4, anchor="ne")

        self._show_hint()
        self.input_text.bind("<FocusIn>", self._clear_hint)
        self.input_text.bind("<FocusOut>", self._restore_hint_if_empty)


        # ───────────────────── BOTONES ─────────────────────

        btn_frame = ttk.Frame(parent)
        btn_frame.pack(pady=6)

        ttk.Button(
            btn_frame,
            text="▶  Compilar",
            command=self.analyze
        ).pack(side=tk.LEFT, padx=6)

        ttk.Button(
            btn_frame,
            text="📤 Exportar Excel",
            command=self.export_to_excel
        ).pack(side=tk.LEFT, padx=6)

        ttk.Button(
            btn_frame,
            text="🗑  Limpiar",
            command=self.clear_all
        ).pack(side=tk.LEFT, padx=6)


        # ───────────────────── NOTEBOOK SECUNDARIO DE SALIDA ─────────────────────
        self.output_notebook = ttk.Notebook(parent)
        self.output_notebook.pack(fill=tk.BOTH, expand=True, padx=12, pady=(4, 12))

        # Subpestaña 1: Análisis léxico-sintáctico
        self.tab_analysis_output = ttk.Frame(self.output_notebook)
        self.output_notebook.add(
            self.tab_analysis_output,
            text="Análisis léxico-sintáctico"
        )

        self.token_text = self._make_output(self.tab_analysis_output)

        copy_analysis_button = self._make_icon_button(
            self.tab_analysis_output,
            "📋",
            lambda: self.copy_widget_text(self.token_text, "análisis"),
            "Copiar análisis"
        )
        copy_analysis_button.place(relx=1.0, x=-28, y=8, anchor="ne")

        # Subpestaña 2: Tabla VLSM
        self.tab_vlsm_output = ttk.Frame(self.output_notebook)
        self.output_notebook.add(
            self.tab_vlsm_output,
            text="Tabla VLSM"
        )
        self.vlsm_text = self._make_output(self.tab_vlsm_output)

        # Subpestaña 3: Errores
        self.tab_errors_output = ttk.Frame(self.output_notebook)
        self.output_notebook.add(
            self.tab_errors_output,
            text="Registro de Errores"
        )

        errors_header = ttk.Frame(self.tab_errors_output)
        errors_header.pack(fill=tk.X, padx=4, pady=(4, 0))

        copy_errors_button = self._make_icon_button(
            errors_header,
            "📋",
            lambda: self.copy_widget_text(self.error_text, "errores"),
            "Copiar errores"
        )
        copy_errors_button.pack(side=tk.RIGHT, padx=4)

        ttk.Button(
            errors_header,
            text="Borrar Errores",
            command=self.clear_errors
        ).pack(side=tk.LEFT, padx=4)

        self.error_text = self._make_output(self.tab_errors_output, fg="#f38ba8")


    # ───────────────────── MÉTODOS AUXILIARES ─────────────────────

    # Crea un área de texto para mostrar resultados.
    # Se usa en análisis, tabla VLSM y errores.
    def _make_output(self, parent, fg="#cdd6f4"):
        t = scrolledtext.ScrolledText(
            parent, wrap=tk.WORD, width=80, height=12,
            font=font.Font(family="Courier New", size=10),
            bg="#181825", fg=fg,
            insertbackground="#f38ba8", state=tk.DISABLED,
            relief="flat", bd=4
        )
        t.pack(fill=tk.BOTH, expand=True, padx=4, pady=4)
        return t

    # Crea un botón pequeño con tooltip.
    # Se usa para pegar, copiar y guardar el árbol.
    def _make_icon_button(self, parent, text, command, tooltip_text):
        button = ttk.Button(
            parent,
            text=text,
            width=3,
            command=command
        )

        Tooltip(button, tooltip_text)

        return button

    # Construye la pestaña de tablas internas.
    # Muestra tabla de tokens y palabras reservadas.
    def _build_tables(self, parent):
        self.tables_notebook = ttk.Notebook(parent)
        self.tables_notebook.pack(fill=tk.BOTH, expand=True, padx=6, pady=6)

        tab_token_table = ttk.Frame(self.tables_notebook)
        self.tables_notebook.add(tab_token_table, text="Tabla de tokens")

        cols_tokens = ("Tipo", "Valor", "Línea", "Posición")

        self.tv_tokens = ttk.Treeview(
            tab_token_table,
            columns=cols_tokens,
            show="headings"
        )

        for col in cols_tokens:
            self.tv_tokens.heading(col, text=col)

            if col == "Tipo":
                self.tv_tokens.column(col, anchor="center", width=280)
            elif col == "Valor":
                self.tv_tokens.column(col, anchor="center", width=220)
            elif col == "Línea":
                self.tv_tokens.column(col, anchor="center", width=100)
            else:
                self.tv_tokens.column(col, anchor="center", width=120)

        self.tv_tokens.tag_configure(
            "separator",
            background="#89b4fa",
            foreground="#1e1e2e",
            font=("Courier New", 10, "bold")
        )

        scroll_tokens_y = ttk.Scrollbar(
            tab_token_table,
            orient=tk.VERTICAL,
            command=self.tv_tokens.yview
        )

        self.tv_tokens.configure(yscrollcommand=scroll_tokens_y.set)

        self.tv_tokens.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(8, 0), pady=8)
        scroll_tokens_y.pack(side=tk.RIGHT, fill=tk.Y, padx=(0, 8), pady=8)

        tab_reserved_table = ttk.Frame(self.tables_notebook)
        self.tables_notebook.add(tab_reserved_table, text="Tabla de palabras reservadas")

        cols_reserved = ("Palabra reservada", "Cantidad")

        self.tv_reserved = ttk.Treeview(
            tab_reserved_table,
            columns=cols_reserved,
            show="headings"
        )

        for col in cols_reserved:
            self.tv_reserved.heading(col, text=col)
            self.tv_reserved.column(col, anchor="center", width=180)

        scroll_reserved_y = ttk.Scrollbar(
            tab_reserved_table,
            orient=tk.VERTICAL,
            command=self.tv_reserved.yview
        )

        self.tv_reserved.configure(yscrollcommand=scroll_reserved_y.set)

        self.tv_reserved.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(8, 0), pady=8)
        scroll_reserved_y.pack(side=tk.RIGHT, fill=tk.Y, padx=(0, 8), pady=8)

    # Construye la pestaña donde se muestran los árboles.
    # Inicializa las listas usadas para guardar árboles generados.
    def _build_tree(self, parent):
        self.tree_notebook = ttk.Notebook(parent)
        self.tree_notebook.pack(fill=tk.BOTH, expand=True, padx=6, pady=6)

        self.tree_canvases = []
        self.tree_names = []
        self.tree_blocks = []

    # Escribe texto en un área de salida deshabilitada.
    # Puede limpiar el contenido anterior antes de escribir.
    def _write(self, widget, text, clear=False):
        widget.config(state=tk.NORMAL)
        if clear:
            widget.delete("1.0", tk.END)
        widget.insert(tk.END, text)
        widget.config(state=tk.DISABLED)

    # Borra el texto de ejemplo al entrar al área de entrada.
    # Solo se ejecuta si el placeholder está activo.
    def _clear_hint(self, event):
        if self._hint_active:
            self.input_text.delete("1.0", tk.END)
            self.input_text.config(fg="#cdd6f4")
            self._hint_active = False
            self.input_text.mark_set("insert", "1.0")
            self.root.after_idle(self.linenumbers.redraw)

    # Restaura el texto de ejemplo si la entrada queda vacía.
    # Se activa cuando el usuario sale del cuadro de entrada.
    def _restore_hint_if_empty(self, event=None):
        if self._hint_active:
            return

        content = self.input_text.get("1.0", tk.END).strip()

        if not content:
            self._show_hint()

    # Muestra el texto de ejemplo al iniciar el programa
    # y cuando el usuario presiona el botón Limpiar.
    def _show_hint(self):
        self.input_text.config(state=tk.NORMAL)
        self.input_text.delete("1.0", tk.END)
        self.input_text.insert("1.0", self.hint_text)
        self.input_text.config(fg="#585b70")
        self._hint_active = True
        self.root.after_idle(self.linenumbers.redraw)

    # ───────────────────── EVENTOS ─────────────────────

    # Actualiza números de línea y resaltado de palabras.
    # Se ejecuta cada vez que cambia la entrada.
    def on_key_release(self, event=None):
        self.linenumbers.redraw()
        self.highlight_reserved_words()

    # Resalta visualmente las palabras reservadas del DSL.
    # Usa expresiones regulares sobre el texto de entrada.
    def highlight_reserved_words(self):
        patterns = {
            "IP": (r"\bIP(?=\s|\d|$)", "#89b4fa"),
            "MASK": (r"\bMASK(?=\s|/|$)", "#a6e3a1"),
            "HOSTS": (r"\bHOSTS(?=\s|\d|,|$)", "#fab387"),
            "NAME": (r"\bNAME\b", "#f9e2af"),
        }

        text = self.input_text.get("1.0", tk.END)

        for word in patterns:
            tag = f"rw_{word}"
            self.input_text.tag_remove(tag, "1.0", tk.END)

        for word, (pattern, color) in patterns.items():
            tag = f"rw_{word}"

            for match in re.finditer(pattern, text):
                start = f"1.0 + {match.start()} chars"
                end = f"1.0 + {match.end()} chars"

                self.input_text.tag_add(tag, start, end)
                self.input_text.tag_config(
                    tag,
                    foreground=color,
                    font=font.Font(
                        family="Courier New",
                        size=10,
                        weight="bold"
                    )
                )


    # ───────────────────── ANÁLISIS ─────────────────────
    # Ejecuta el flujo principal:
    # 1. Lee la entrada del usuario.
    # 2. Realiza el análisis léxico.
    # 3. Realiza el análisis sintáctico.
    # 4. Calcula la tabla VLSM si no hay errores.
    # 5. Muestra los resultados en la interfaz.
    def analyze(self):
        if self._hint_active:
            messagebox.showwarning("Advertencia", "Introduce un código para analizar.")
            return

        code = self.input_text.get("1.0", tk.END).strip()

        if not code:
            messagebox.showwarning("Advertencia", "El área de entrada está vacía.")
            return

        # Siempre regresar visualmente a Entrada / Salida al analizar
        self.notebook.select(self.tab_io)

        self.clear_errors()
        self._write(self.token_text, "", clear=True)
        self._write(self.vlsm_text, "", clear=True)

        self.vlsm_data = None
        self.tokens = []
        self.draw_tree([])

        # ── 1. ANÁLISIS LÉXICO ───────────────────────────────
        lexer = VLSMLexer()
        tokens, lex_errors = lexer.tokenize(code)


        analysis_out = "=== ANÁLISIS LÉXICO ===\n"

        if tokens:
            for token in tokens:
                token_type, value, line, col = token
                analysis_out += (
                    f"Token: {token_type:<12} Valor: {value:<18} "
                    f"(Línea {line}, Posición {col})\n"
                )
        else:
            analysis_out += "No se reconocieron tokens.\n"

        self._write(self.token_text, analysis_out, clear=True)

        # Aquí se guardan errores léxicos, pero NO se detiene el análisis todavía.
        error_out = ""

        if lex_errors:
            error_out += "=== ERRORES LÉXICOS ===\n"

            for i, error in enumerate(lex_errors, start=1):
                error_out += f"  {i}. {error}\n"

            error_out += "\n"

        # ── 2. ANÁLISIS SINTÁCTICO ───────────────────────────
        # Aunque haya errores léxicos, se ejecuta el parser con los tokens reconocidos.
        # Esto permite detectar errores sintácticos en otras líneas.
        parser = VLSMParser(tokens)
        blocks = parser.parse()
        syntax_errors = parser.errors

        self.valid_blocks = blocks

        valid_lines = {block["source_line"] for block in blocks}
        valid_tokens = [token for token in tokens if token[2] in valid_lines]

        self.tokens = valid_tokens
        self.populate_tables()

        analysis_out += "\n=== ANÁLISIS SINTÁCTICO ===\n"

        if syntax_errors:
            error_out += "=== ERRORES SINTÁCTICOS ===\n"

            for i, error in enumerate(syntax_errors, start=1):
                error_out += f"  {i}. {error}\n"

            error_out += "\n"

        # ── 3. SI HAY ERRORES, MOSTRAR TODOS Y DETENER CÁLCULO ──
        has_errors = bool(lex_errors or syntax_errors)

        if has_errors and not blocks:
            analysis_out += "Se encontraron errores y no hay bloques válidos para procesar.\n"
            analysis_out += f"\nErrores léxicos: {len(lex_errors)}\n"
            analysis_out += f"Errores sintácticos: {len(syntax_errors)}\n"
            analysis_out += f"Bloques válidos reconocidos: {len(blocks)}\n"

            self._write(self.token_text, analysis_out, clear=True)
            self._write(self.error_text, error_out, clear=True)

            self.draw_tree([])
            self.vlsm_data = None
            self._write(
                self.vlsm_text,
                "No se generó tabla VLSM porque no existen bloques válidos.\n",
                clear=True
            )

            self.notebook.select(self.tab_io)
            self.output_notebook.select(self.tab_errors_output)
            return

        if has_errors and blocks:
            analysis_out += "Se encontraron errores, pero se procesarán únicamente los bloques válidos.\n"
            analysis_out += f"\nErrores léxicos: {len(lex_errors)}\n"
            analysis_out += f"Errores sintácticos: {len(syntax_errors)}\n"
            analysis_out += f"Bloques válidos reconocidos: {len(blocks)}\n\n"

            self._write(self.error_text, error_out, clear=True)

        # ── 4. SI NO HAY BLOQUES VÁLIDOS ─────────────────────
        if not blocks:
            analysis_out += "No se encontraron bloques válidos.\n"

            self._write(self.token_text, analysis_out, clear=True)
            self.output_notebook.select(self.tab_analysis_output)
            return

        # ── 5. MOSTRAR BLOQUES VÁLIDOS ───────────────────────
        for i, block in enumerate(blocks, start=1):
            nombre = block.get("name") if block.get("name") else "Sin nombre"

            analysis_out += (
                f"Red #{i}:\n"
                f"  IP: {block['ip_address']}\n"
                f"  Máscara: {block['subnet_mask']}\n"
                f"  Hosts requeridos: {block['num_hosts']}\n"
                f"  Nombre: {nombre}\n"
            )

        # ── 6. ÁRBOL SINTÁCTICO ──────────────────────────────
        tree_parser = VLSMParser(valid_tokens)
        tree_blocks = tree_parser.parse_with_tree()
        self.draw_tree(tree_blocks)

        # ── 7. CÁLCULO VLSM ─────────────────────────────────
        all_results = []

        for block in blocks:
            try:
                result = calculate_vlsm(
                    block["ip_address"],
                    block["subnet_mask"],
                    block["num_hosts"],
                    nombre_red=block.get("name")
                )

                all_results.extend(result)

            except Exception as ex:
                self._write(
                    self.error_text,
                    f"Error al calcular VLSM para {block['ip_address']}: {ex}\n",
                    clear=True
                )

                self.draw_tree([])
                self.vlsm_data = None

                self.notebook.select(self.tab_io)
                self.output_notebook.select(self.tab_errors_output)
                return

        self.vlsm_data = all_results

        # ── 8. SALIDA DETALLADA TIPO COMPILADOR ───────
        analysis_out += "\n=== RESULTADOS VLSM ===\n"

        for result in all_results:
            analysis_out += (
                f"Hosts solicitados: {result['hosts_solicitados']}\n"
                f"Hosts disponibles: {result['hosts_disponibles']}\n"
                f"Dirección de red: {result['direccion_de_red']}\n"
                f"Nueva máscara: {result['nueva_mascara']}\n"
                f"Máscara decimal: {result['mascara_decimal']}\n"
                f"Primera IP utilizable: {result['primera_ip_utilizable']}\n"
                f"Última IP utilizable: {result['ultima_ip_utilizable']}\n"
                f"Dirección de broadcast: {result['direccion_de_broadcast']}\n"
                "-----------------------------------------------------------\n"
            )
        
        total_errores = len(lex_errors) + len(syntax_errors)

        analysis_out += "\n=== RESUMEN DEL ANÁLISIS ===\n"
        analysis_out += f"Tokens reconocidos: {len(tokens)}\n"
        analysis_out += f"Bloques válidos: {len(blocks)}\n"
        analysis_out += f"Subredes generadas: {len(all_results)}\n"
        analysis_out += f"Errores encontrados: {total_errores}\n"


        self._write(self.token_text, analysis_out, clear=True)

        # ── 9. TABLA VLSM FORMATEADA ────────────────────────
        vlsm_out = "=== TABLA VLSM ===\n\n"

        grouped = {}

        for result in all_results:
            key = result.get("nombre_red") or result["ip_base"]
            grouped.setdefault(key, []).append(result)

        for nombre, subredes in grouped.items():
            vlsm_out += f"  Red: {nombre}\n"

            header = (
                f"  {'#':<4} {'Hosts Solic.':<14} {'Hosts Disp.':<13} "
                f"{'Dirección Red':<18} {'CIDR':<8} {'Máscara':<18} "
                f"{'Primera IP':<16} {'Última IP':<16} {'Broadcast'}\n"
            )

            separator = "  " + "─" * (len(header) + 3) + "\n"

            vlsm_out += separator
            vlsm_out += header
            vlsm_out += separator

            for i, subred in enumerate(subredes, start=1):
                vlsm_out += (
                    f"  {i:<4} "
                    f"{subred['hosts_solicitados']:<14} "
                    f"{subred['hosts_disponibles']:<13} "
                    f"{subred['direccion_de_red']:<18} "
                    f"{subred['nueva_mascara']:<8} "
                    f"{subred['mascara_decimal']:<18} "
                    f"{subred['primera_ip_utilizable']:<16} "
                    f"{subred['ultima_ip_utilizable']:<16} "
                    f"{subred['direccion_de_broadcast']}\n"
                )

            vlsm_out += "\n"

        self._write(self.vlsm_text, vlsm_out, clear=True)

        # Si todo salió bien, mostrar directamente la tabla VLSM
        if has_errors:
            self.notebook.select(self.tab_io)
            self.output_notebook.select(self.tab_errors_output)

            messagebox.showwarning(
                "Análisis terminado con errores",
                f"Se encontraron errores, pero se calcularon {len(all_results)} subred(es) de bloques válidos."
            )
        else:
            self.notebook.select(self.tab_io)
            self.output_notebook.select(self.tab_vlsm_output)

            messagebox.showinfo(
                "Análisis completo",
                f"✓ {len(all_results)} subred(es) calculada(s) correctamente."
    )


    # ───────────────────── LIMPIAR ─────────────────────
    
    # Limpia únicamente el área de errores.
    # Se usa antes de analizar o al presionar Borrar Errores.
    def clear_errors(self):
        self._write(self.error_text, "", clear=True)

    # Limpia entrada, salidas, tablas y árbol.
    # Después restaura el texto de ejemplo inicial.
    def clear_all(self):
        self.clear_errors()

        self._write(self.token_text, "", clear=True)
        self._write(self.vlsm_text, "", clear=True)

        self.vlsm_data = None
        self.tokens = []
        self.valid_blocks = []

        for item in self.tv_tokens.get_children():
            self.tv_tokens.delete(item)

        for item in self.tv_reserved.get_children():
            self.tv_reserved.delete(item)

        self.draw_tree([])

        self._show_hint()

        # Regresar a la vista principal
        self.notebook.select(self.tab_io)
        self.output_notebook.select(self.tab_analysis_output)

    # ───────────────────── EXPORTAR A EXCEL ─────────────────────
    # Exporta la tabla VLSM actual a Excel.
    # Muestra error si todavía no hay resultados generados.
    def export_to_excel(self):
        if self.vlsm_data:
            export_to_excel(self.vlsm_data)
        else:
            messagebox.showerror("Error", "No hay datos para exportar.")


    # ───────────────────── TABLAS INTERNAS ─────────────────────
    # Llena las tablas internas de la interfaz.
    # Agrupa tokens válidos usando filas separadoras por línea/bloque.
    def populate_tables(self):
        for item in self.tv_tokens.get_children():
            self.tv_tokens.delete(item)

        for item in self.tv_reserved.get_children():
            self.tv_reserved.delete(item)

        # Crear etiquetas para cada bloque válido.
        block_labels = {}

        for index, block in enumerate(getattr(self, "valid_blocks", []), start=1):
            line = block.get("source_line")
            name = block.get("name") if block.get("name") else "Sin nombre"
            ip = block.get("ip_address", "")

            block_labels[line] = f"──── Línea {line} - Red #{index} ────"

        # Agrupar tokens por línea.
        tokens_by_line = {}

        for token in self.tokens:
            token_type, value, line, col = token
            tokens_by_line.setdefault(line, []).append(token)

        # Insertar una fila separadora por cada línea válida.
        for line in sorted(tokens_by_line.keys()):
            separator_text = block_labels.get(line, f"──── # {line} ────")

            self.tv_tokens.insert(
                "",
                tk.END,
                values=(separator_text, "", "", ""),
                tags=("separator",)
            )

            for token in tokens_by_line[line]:
                token_type, value, line, col = token

                self.tv_tokens.insert(
                    "",
                    tk.END,
                    values=(token_type, value, line, col)
                )

        # Tabla de palabras reservadas solo con tokens válidos.
        reserved_words = ["IP", "MASK", "HOSTS", "NAME"]
        counts = Counter(token[0] for token in self.tokens if token[0] in reserved_words)

        for word in reserved_words:
            if counts[word] > 0:
                self.tv_reserved.insert("", tk.END, values=(word, counts[word]))


    # ───────────────────── ÁRBOL SINTÁCTICO ─────────────────────

    # Dibuja el árbol sintáctico dentro de la interfaz.
    # Crea una pestaña por cada bloque válido analizado.
    def draw_tree(self, tree):
        # Elimina las pestañas de árboles anteriores antes de dibujar uno nuevo.
        for tab in self.tree_tabs:
            self.tree_notebook.forget(tab)

        # Reinicia las listas que guardan referencias a los árboles actuales.
        self.tree_tabs = []
        self.tree_canvases = []
        self.tree_names = []
        self.tree_blocks = []

        # Si no hay árboles válidos, no dibuja nada y regresa.
        if not tree:
            return

        # Configuraciones visuales para el dibujo del árbol.
        V_SPACING = 85
        H_SPACING = 120

        # Fuentes y colores para nodos, hojas y fondo del árbol.
        node_font = ("Courier New", 10, "bold")
        leaf_font = ("Courier New", 9)

        # Colores inspirados en el tema "Catppuccin" para una apariencia moderna y agradable.
        bg_canvas = "#181825"
        line_color = "#89b4fa"
        node_bg = "#313244"
        root_bg = "#45475a"
        leaf_bg = "#1e1e2e"
        text_color = "#cdd6f4"
        border_color = "#89b4fa"

        # Funciones auxiliares para manejar la estructura del árbol.
        # Se asume que el árbol se representa como tuplas (label, children) donde children es una lista de nodos hijos.
        def is_branch(node):
            return (
                isinstance(node, tuple)
                and len(node) == 2
                and isinstance(node[1], list)
            )

        # Obtiene la etiqueta de un nodo para mostrar en el dibujo.
        # Si el nodo es una rama, muestra su etiqueta principal.
        def node_label(node):
            if is_branch(node):
                return str(node[0])

            if isinstance(node, tuple) and len(node) == 2:
                if node[0] == "VALOR":
                    return str(node[1])
                return f"{node[0]}: {node[1]}"

            return str(node)

        # Obtiene los hijos de un nodo si es una rama, o una lista vacía si es una hoja.
        # Esto permite recorrer el árbol recursivamente para calcular posiciones y dibujar.
        def node_children(node):
            if is_branch(node):
                return node[1]
            return []

        # Calcula el ancho total que ocupa un subárbol para posicionar los nodos correctamente.
        # El ancho mínimo es H_SPACING, pero si el subárbol tiene ramas anchas, se ajusta en consecuencia.
        def subtree_width(node):
            children = node_children(node)

            if not children:
                return H_SPACING
            # El ancho del subárbol es la suma de los anchos de sus hijos, o al menos H_SPACING para evitar que se superpongan.
            return max(
                H_SPACING,
                sum(subtree_width(child) for child in children)
            )

        # Calcula la altura total del árbol para determinar el tamaño vertical del canvas.
        # La altura se mide en niveles, y cada nivel se separa por V_SPACING
        def tree_height(node, depth=1):
            children = node_children(node)

            if not children:
                return depth

            return max(tree_height(child, depth + 1) for child in children)

        # Dibuja un nodo y sus conexiones a los hijos en el canvas.
        # Se llama recursivamente para cada nodo del árbol, posicionando los hijos debajo del padre.
        def draw_node(canvas, node, x, y, is_root=False):
            children = node_children(node)
            label = node_label(node)

            if children:
                widths = [subtree_width(child) for child in children]
                total_width = sum(widths)
                start_x = x - total_width / 2
                child_positions = []

                for width, child in zip(widths, children):
                    child_x = start_x + width / 2
                    child_positions.append((child, child_x))
                    start_x += width

                child_y = y + V_SPACING

                for child, child_x in child_positions:
                    canvas.create_line(
                        x, y + 22,
                        child_x, child_y - 22,
                        fill=line_color,
                        width=2
                    )
                    draw_node(canvas, child, child_x, child_y)

            font_used = node_font if children else leaf_font
            fill_used = root_bg if is_root else (node_bg if children else leaf_bg)

            temp_text = canvas.create_text(
                x,
                y,
                text=label,
                font=font_used
            )

            bbox = canvas.bbox(temp_text)
            canvas.delete(temp_text)

            pad_x = 18
            pad_y = 10

            x0 = bbox[0] - pad_x
            y0 = bbox[1] - pad_y
            x1 = bbox[2] + pad_x
            y1 = bbox[3] + pad_y

            canvas.create_oval(
                x0, y0, x1, y1,
                fill=fill_used,
                outline=border_color,
                width=2
            )

            canvas.create_text(
                x,
                y,
                text=label,
                fill=text_color,
                font=font_used
            )

        # Para cada bloque válido en el árbol sintáctico, crea una pestaña nueva y dibuja el árbol correspondiente.
        # Calcula el tamaño necesario del canvas para acomodar el árbol completo con sus ramas y hojas.
        for index, block in enumerate(tree, start=1):
            tab = ttk.Frame(self.tree_notebook)
            self.tree_tabs.append(tab)
            self.tree_notebook.add(tab, text=f"Árbol {index}")

            width = max(900, subtree_width(block) + 200)
            height = max(500, tree_height(block) * V_SPACING + 120)

            canvas = tk.Canvas(
                tab,
                bg=bg_canvas,
                scrollregion=(0, 0, width, height)
            )

            tree_name = self.get_tree_name(block, index)
            self.tree_names.append(tree_name)
            self.tree_canvases.append(canvas)
            self.tree_blocks.append(block)

            x_scroll = ttk.Scrollbar(tab, orient=tk.HORIZONTAL, command=canvas.xview)
            y_scroll = ttk.Scrollbar(tab, orient=tk.VERTICAL, command=canvas.yview)

            canvas.configure(
                xscrollcommand=x_scroll.set,
                yscrollcommand=y_scroll.set
            )

            canvas.grid(row=0, column=0, sticky="nsew")
            y_scroll.grid(row=0, column=1, sticky="ns")
            x_scroll.grid(row=1, column=0, sticky="ew")

            tab.grid_rowconfigure(0, weight=1)
            tab.grid_columnconfigure(0, weight=1)

            save_tree_button = self._make_icon_button(
                tab,
                "💾",
                self.save_current_tree_image,
                "Guardar árbol como PNG"
            )
            save_tree_button.place(relx=1.0, x=-23, y=8, anchor="ne")
            save_tree_button.lift()

            draw_node(canvas, block, width / 2, 60, is_root=True)

    # Obtiene un nombre seguro para identificar el árbol.
    # Usa el nombre de red o la IP si no hay identificador.
    def get_tree_name(self, block, index):
        root_label = str(block[0]) if isinstance(block, tuple) else ""

        if ":" in root_label:
            possible_name = root_label.split(":", 1)[1].strip()

            if possible_name:
                return clean_name_for_file(possible_name, f"Red_{index}")

        # Si no hay un nombre explícito, intenta extraer la IP para usarla como identificador del árbol.
        # Busca recursivamente en el árbol sintáctico un nodo con etiqueta "IP_ADDRESS" y obtiene su valor.
        def search_ip(node):
            if not isinstance(node, tuple) or len(node) != 2:
                return None

            label, children = node

            if label == "IP_ADDRESS":
                if isinstance(children, list) and children:
                    first_child = children[0]

                    if isinstance(first_child, tuple) and len(first_child) == 2:
                        return str(first_child[1])

                return None

            if isinstance(children, list):
                for child in children:
                    result = search_ip(child)

                    if result:
                        return result

            return None

        ip_found = search_ip(block)

        if ip_found:
            return clean_name_for_file(ip_found, f"Red_{index}")

        return f"Red_{index}"

    # Genera una imagen completa del árbol usando Pillow.
    # No depende del scroll ni del tamaño visible del canvas.
    def render_tree_to_image(self, block):
        V_SPACING = 120
        H_SPACING = 180

        BG_CANVAS = "#181825"
        LINE_COLOR = "#89b4fa"
        NODE_BG = "#313244"
        ROOT_BG = "#45475a"
        LEAF_BG = "#1e1e2e"
        TEXT_COLOR = "#cdd6f4"
        BORDER_COLOR = "#89b4fa"

        try:
            node_font = ImageFont.truetype("consola.ttf", 18)
            leaf_font = ImageFont.truetype("consola.ttf", 16)
        except Exception:
            node_font = ImageFont.load_default()
            leaf_font = ImageFont.load_default()

        # Funciones auxiliares para manejar la estructura del árbol.
        # Se asume que el árbol se representa como tuplas (label, children) 
        # donde children es una lista de nodos hijos.
        def is_branch(node):
            return (
                isinstance(node, tuple)
                and len(node) == 2
                and isinstance(node[1], list)
            )

        # Obtiene la etiqueta de un nodo para mostrar en el dibujo.
        # Si el nodo es una rama, muestra su etiqueta principal.
        def node_label(node):
            if is_branch(node):
                return str(node[0])

            if isinstance(node, tuple) and len(node) == 2:
                if node[0] == "VALOR":
                    return str(node[1])
                return f"{node[0]}: {node[1]}"

            return str(node)

        # Obtiene los hijos de un nodo si es una rama, o una lista vacía si es una hoja.
        # Esto permite recorrer el árbol recursivamente para calcular posiciones y dibujar.
        def node_children(node):
            if is_branch(node):
                return node[1]
            return []

        # Calcula el ancho total que ocupa un subárbol para posicionar los nodos correctamente.
        # El ancho mínimo es H_SPACING, pero si el subárbol tiene ramas "an"
        def subtree_width(node):
            children = node_children(node)

            if not children:
                return H_SPACING

            return max(
                H_SPACING,
                sum(subtree_width(child) for child in children)
            )

        # Calcula la altura total del árbol para determinar el tamaño vertical del canvas.
        # La altura se mide en niveles, y cada nivel se separa por V_SPACING
        def tree_height(node, depth=1):
            children = node_children(node)

            if not children:
                return depth

            return max(tree_height(child, depth + 1) for child in children)

        width = max(1200, subtree_width(block) + 260)
        height = max(700, tree_height(block) * V_SPACING + 180)

        image = Image.new("RGB", (width, height), BG_CANVAS)
        draw = ImageDraw.Draw(image)

        # Función auxiliar para medir el tamaño del texto con la fuente dada.
        # Esto es necesario para centrar correctamente los nodos y ajustar el tamaño de las elipses.
        def text_size(text, font_used):
            bbox = draw.textbbox((0, 0), text, font=font_used)
            return bbox[2] - bbox[0], bbox[3] - bbox[1]

        # Dibuja un nodo y sus conexiones a los hijos en la imagen.
        # Se llama recursivamente para cada nodo del árbol, posicionando los hijos debajo del padre.
        def draw_node(node, x, y, is_root=False):
            children = node_children(node)
            label = node_label(node)

            if children:
                widths = [subtree_width(child) for child in children]
                total_width = sum(widths)
                start_x = x - total_width / 2
                child_positions = []

                for width_child, child in zip(widths, children):
                    child_x = start_x + width_child / 2
                    child_positions.append((child, child_x))
                    start_x += width_child

                child_y = y + V_SPACING

                for child, child_x in child_positions:
                    draw.line(
                        [(x, y + 28), (child_x, child_y - 28)],
                        fill=LINE_COLOR,
                        width=3
                    )
                    draw_node(child, child_x, child_y)

            font_used = node_font if children else leaf_font
            fill_used = ROOT_BG if is_root else (NODE_BG if children else LEAF_BG)

            text_w, text_h = text_size(label, font_used)

            pad_x = 24
            pad_y = 14

            x0 = int(x - text_w / 2 - pad_x)
            y0 = int(y - text_h / 2 - pad_y)
            x1 = int(x + text_w / 2 + pad_x)
            y1 = int(y + text_h / 2 + pad_y)

            draw.ellipse(
                [x0, y0, x1, y1],
                fill=fill_used,
                outline=BORDER_COLOR,
                width=3
            )

            draw.text(
                (int(x - text_w / 2), int(y - text_h / 2)),
                label,
                fill=TEXT_COLOR,
                font=font_used
            )

        # Para cada bloque válido en el árbol sintáctico, crea una pestaña nueva y dibuja el árbol correspondiente.
        draw_node(block, width / 2, 80, is_root=True)

        return image

    # Guarda el árbol seleccionado como imagen PNG.
    # Usa un nombre automático basado en la red.
    def save_current_tree_image(self):
        if not hasattr(self, "tree_blocks") or not self.tree_blocks:
            messagebox.showerror("Error", "No hay árbol para guardar.")
            return

        try:
            selected_tab = self.tree_notebook.select()

            if not selected_tab:
                messagebox.showerror("Error", "No hay árbol seleccionado.")
                return

            index = self.tree_notebook.index(selected_tab)
            block = self.tree_blocks[index]

            tree_name = "Red"

            if hasattr(self, "tree_names") and index < len(self.tree_names):
                tree_name = self.tree_names[index]

        except Exception:
            messagebox.showerror("Error", "No se pudo obtener el árbol seleccionado.")
            return

        nombre_archivo = f"Arbol_{index + 1}_{tree_name}.png"

        file_path = filedialog.asksaveasfilename(
            initialfile=nombre_archivo,
            defaultextension=".png",
            filetypes=[("Imagen PNG", "*.png")]
        )

        if not file_path:
            return

        try:
            image = self.render_tree_to_image(block)
            image.save(file_path)
            messagebox.showinfo("Guardar árbol", "Árbol guardado correctamente.")
        except Exception as ex:
            messagebox.showerror("Error", f"No se pudo guardar el árbol: {ex}")


    # ───────────────────── PORTAPAPELES ─────────────────────    

    # Pega en la entrada el texto actual del portapapeles.
    # También actualiza resaltado y números de línea.
    def paste_clipboard_to_input(self):
        try:
            clipboard_text = self.root.clipboard_get()
        except tk.TclError:
            messagebox.showwarning("Advertencia", "No hay texto en el portapapeles.")
            return

        clipboard_text = clipboard_text.strip()

        if not clipboard_text:
            messagebox.showwarning("Advertencia", "El portapapeles está vacío.")
            return

        self.input_text.config(state=tk.NORMAL)
        self.input_text.delete("1.0", tk.END)
        self.input_text.insert("1.0", clipboard_text)
        self.input_text.config(fg="#cdd6f4")

        self._hint_active = False
        self.root.after_idle(self.linenumbers.redraw)
        self.highlight_reserved_words()

        self.notebook.select(self.tab_io)
        self.output_notebook.select(self.tab_analysis_output)

    # Copia al portapapeles el contenido de un área de salida.
    # Se usa para copiar análisis o errores.
    def copy_widget_text(self, widget, nombre):
        text = widget.get("1.0", tk.END).strip()

        if not text:
            messagebox.showwarning("Advertencia", f"No hay {nombre} para copiar.")
            return

        self.root.clipboard_clear()
        self.root.clipboard_append(text)
        self.root.update()

        messagebox.showinfo("Copiar", f"Contenido de {nombre} copiado al portapapeles.")   


# Punto de entrada del programa.
# Este bloque solo se ejecuta cuando el archivo se abre directamente,
# no cuando se importa desde otro archivo.
if __name__ == "__main__": # Inicia la aplicación gráfica.
    root = tk.Tk() # Crea la ventana principal de la interfaz gráfica.
    root.minsize(800, 600) # Define el tamaño mínimo de la ventana.
    app = VLSMApp(root) # Crea la aplicación TecNet Finder dentro de la ventana principal.
    root.mainloop() # Mantiene abierta la ventana y espera las acciones del usuario.
